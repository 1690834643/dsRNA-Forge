#!/usr/bin/env python3
"""
Delivery-readiness regression tests.

These tests cover behavior that matters for an end-user Windows build, not just
developer smoke tests.
"""

from dsforge.core.offtarget import OffTargetScreener
from dsforge.core.offtarget_index import OffTargetRiskIndex
from pathlib import Path
from dsforge.core.redundancy import cluster_redundant_results
from dsforge.core.sgrna import (
    design_sgrna_cloning_oligos,
    scan_sgrna_candidates,
    score_sgrna_offtargets,
)
from dsforge.core.sequence import (
    TranscriptomeIndex,
    clone_with_custom_sequence,
    merge_background_transcriptomes,
    generate_candidates,
    load_first_fasta_record,
    normalize_sequence,
)
from dsforge.core.diagnostics import diagnose_design_outcome
from dsforge.core.explain import explain_result, render_region_map
from dsforge.core.primers import T7_PROMOTER, calculate_primer_tm, check_primer_specificity, design_t7_primers
from dsforge.core.presets import apply_preset_to_config
from dsforge.core.project import load_project_file, save_project_file
from dsforge.core.thermodynamics import ThermodynamicsCalculator
from dsforge.core.validation import build_validation_hits
from dsforge.controller.design_task_parallel import _score_long_dsRNA_batch, _score_sirna_batch
from dsforge.controller.design_task import DesignCancelled, DesignConfig, DesignTask
from dsforge.controller.design_task_parallel import ParallelDesignTask
from dsforge.controller.exporter import ResultExporter
from dsforge.database.manager import DatabaseManager
from dsforge.database.schema import init_database
from dsforge.gui.workers import create_design_task


def test_offtarget_excludes_intended_target_sequence():
    index = TranscriptomeIndex()
    index.sequences = {
        "target_gene": "CCAUGCAUGCAUGCAUGCAUGCAUGAA",
        "unrelated_gene": "UUUUUUUUUUUUUUUUUUUUUUUUUUUU",
    }

    screener = OffTargetScreener(index)
    result = screener.screen_sequence(
        "AUGCAUGCAUGCAUGCAUGCA",
        exclude_ids={"target_gene"},
        use_vienna=False,
        seed_7nt=False,
    )

    assert result["passed"] is True, result
    assert result["summary"]["level_1_16bp_hits"] == 0, result


def test_parallel_worker_loads_builtin_scoring_rules():
    batch = [{"sequence": "GCAUGCAUGCAUGCAUGCAUG", "start": 0, "end": 21}]
    results = _score_sirna_batch({
        "batch": batch,
        "target_seq": "GCAUGCAUGCAUGCAUGCAUG",
        "target_seq_id": "target_gene",
        "enabled_rules": ["consensus", "reynolds"],
        "off_target_config": {"level_1_16bp": True},
        "transcriptome_seqs": {"target_gene": "GCAUGCAUGCAUGCAUGCAUG"},
    })

    assert results[0]["consensus_score"] > 0, results[0]
    assert "not found in registry" not in str(results[0]["rules"]), results[0]


def test_gui_worker_uses_parallel_task_when_multiple_cores_requested():
    parallel = create_design_task(None, DesignConfig(n_cores=4))
    single = create_design_task(None, DesignConfig(n_cores=1))

    assert isinstance(parallel, ParallelDesignTask), type(parallel)
    assert isinstance(single, DesignTask), type(single)


def test_load_first_target_fasta_record_normalizes_sequence(tmp_path):
    fasta = tmp_path / "target.fa"
    fasta.write_text(">my target gene\natgc atgc\nNNNN\n", encoding="utf-8")

    record_id, sequence = load_first_fasta_record(str(fasta), fallback_id="custom_target")

    assert record_id == "my_target_gene"
    assert sequence == "AUGCAUGCNNNN"
    assert normalize_sequence(" atgc\nu ") == "AUGCU"


def test_phase16_load_fasta_rejects_empty_or_headerless_files(tmp_path):
    empty = tmp_path / "empty.fa"
    headerless = tmp_path / "headerless.fa"
    empty.write_text("", encoding="utf-8")
    headerless.write_text("AUGCAUGCAUGC\n", encoding="utf-8")

    for path in [empty, headerless]:
        try:
            TranscriptomeIndex(cache_dir=tmp_path / "cache").load_fasta(str(path), use_cache=False)
            assert False, f"Expected ValueError for {path}"
        except ValueError as exc:
            assert "No FASTA records found" in str(exc)


def test_phase16_generate_candidates_skips_ambiguous_windows_but_keeps_clean_flanks():
    seq = "AUGCAUGCAUNNNNNGCAUGCAUGCA"

    candidates = list(generate_candidates(seq, "siRNA", 5, 5, gc_min=0, gc_max=100, exclude_poly=9999))

    assert candidates, candidates
    assert all("N" not in item["sequence"] for item in candidates), candidates
    assert any(item["start"] == 0 for item in candidates), candidates
    assert any(item["start"] > seq.index("N") for item in candidates), candidates


def test_phase17_ambiguous_iupac_windows_are_reported_in_diagnosis():
    config = DesignConfig(mode="siRNA")
    diagnosis = diagnose_design_outcome(
        target_seq="AUGCAUGCAUNNNNNNNNNNAUGC",
        config=config,
        total_candidates=0,
        passed_candidates=0,
    )

    text = "\n".join(diagnosis["reasons"] + diagnosis["suggestions"])
    assert "N/模糊" in text or "IUPAC" in text, text


def test_no_result_diagnosis_explains_short_target():
    config = DesignConfig(mode="long_dsRNA", length_min=200, length_max=500)
    diagnosis = diagnose_design_outcome(
        target_seq="AUGCAUGCAUGC",
        config=config,
        total_candidates=0,
        passed_candidates=0,
    )

    text = "\n".join(diagnosis["reasons"] + diagnosis["suggestions"])
    assert "目标序列太短" in text
    assert "换用 siRNA/DsiRNA" in text


def test_phase15_sgrna_no_result_diagnosis_explains_missing_spcas9_pam():
    config = DesignConfig(mode="sgRNA")
    diagnosis = diagnose_design_outcome(
        target_seq="AUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAUAU",
        config=config,
        total_candidates=0,
        passed_candidates=0,
    )

    text = "\n".join(diagnosis["reasons"] + diagnosis["suggestions"])
    assert "SpCas9 PAM" in text, text
    assert "NGG" in text and "CCN" in text, text
    assert "目标序列太短" not in text, text


def test_clone_with_custom_sequence_keeps_background_and_adds_target():
    index = TranscriptomeIndex()
    index.sequences = {"background": "UUUUUUUUUUUUUUUUUUUU"}
    index._compute_stats()

    cloned = clone_with_custom_sequence(index, "pasted target", "atgcatgc")

    assert "background" in cloned.sequences
    assert cloned.sequences["pasted_target"] == "AUGCAUGC"
    assert index.sequences == {"background": "UUUUUUUUUUUUUUUUUUUU"}


def test_transcriptome_cache_manifest_loads_saved_without_original_file(tmp_path):
    cache_dir = tmp_path / "cache"
    fasta = tmp_path / "saved_species.fa"
    fasta.write_text(">gene_a\nAUGCAUGCAUGCAUGCAUGCAUGCAUGC\n>gene_b\nUUUUCCCCAAAAGGGG\n", encoding="utf-8")

    index = TranscriptomeIndex(cache_dir=cache_dir).load_fasta(str(fasta))
    saved = TranscriptomeIndex.list_saved(cache_dir=cache_dir)
    assert len(saved) == 1, saved
    assert saved[0]["name"] == "saved_species"

    fasta.unlink()
    reloaded = TranscriptomeIndex.load_saved(saved[0]["key"], cache_dir=cache_dir)

    assert reloaded.sequences == index.sequences
    assert reloaded.get_stats()["num_sequences"] == 2


def test_phase16_transcriptome_and_offtarget_caches_are_json_not_pickle(tmp_path):
    cache_dir = tmp_path / "cache"
    fasta = tmp_path / "species.fa"
    fasta.write_text(">gene_a\nAUGCAUGCAUGCAUGCAUGCAUGC\n>gene_b\nUUUUCCCCAAAAGGGG\n", encoding="utf-8")

    index = TranscriptomeIndex(cache_dir=cache_dir).load_fasta(str(fasta))
    risk_index = OffTargetRiskIndex.from_transcriptome(index, cache_dir=cache_dir)

    transcriptome_cache = Path(index.cache_path)
    risk_cache = cache_dir / f"offtarget_{index.source_hash}_v3_7_16_20_27.json"

    assert transcriptome_cache.read_text(encoding="utf-8").lstrip().startswith("{")
    assert risk_cache.read_text(encoding="utf-8").lstrip().startswith("{")
    assert risk_index.assess_sequence("AUGCAUGCAUGCAUGCAUGC", exclude_ids={"gene_a"})["risk_level"] in {"low", "medium", "high"}


def test_offtarget_risk_summary_ranks_targets_and_suggests_validation():
    index = TranscriptomeIndex()
    query = "AUGCAUGCAUGCAUGCAUGCA"
    index.sequences = {
        "target_gene": "GGGG" + query + "CCCC",
        "danger_gene": "AAAA" + query + "UUUU",
        "seed_gene": "CCCC" + query[1:8] + "GGGG",
    }
    index._compute_stats()

    result = OffTargetScreener(index).screen_sequence(
        query,
        exclude_ids={"target_gene"},
        use_vienna=False,
    )

    assert result["risk_level"] == "high", result
    assert result["risk_score"] >= 80, result
    assert result["top_targets"][0]["target_id"] == "danger_gene", result
    assert "验证" in result["validation_direction"], result


def test_phase14_offtarget_rule_switches_are_honored():
    index = TranscriptomeIndex()
    query = "AUGCAUGCAUGCAUGCAUGCA"
    index.sequences = {
        "target_gene": "GGGG" + query + "CCCC",
        "danger_gene": "AAAA" + query + "UUUU",
    }
    index._compute_stats()

    relaxed = OffTargetScreener(index).screen_sequence(
        query,
        exclude_ids={"target_gene"},
        level_1_16bp=False,
        level_2_20bp=False,
        level_3_27bp=False,
        seed_7nt=False,
        use_vienna=False,
    )
    strict = OffTargetScreener(index).screen_sequence(
        query,
        exclude_ids={"target_gene"},
        level_1_16bp=True,
        level_2_20bp=True,
        seed_7nt=True,
        use_vienna=False,
    )

    assert relaxed["risk_level"] == "low", relaxed
    assert relaxed["risk_score"] == 0, relaxed
    assert relaxed["summary"]["level_2_20bp_hits"] == 0, relaxed
    assert strict["risk_level"] == "high", strict
    assert strict["summary"]["level_2_20bp_hits"] == 1, strict


def test_phase14_parallel_worker_honors_offtarget_config():
    query = "GCAUGCAUGCAUGCAUGCAUG"
    batch = [{"sequence": query, "start": 0, "end": len(query)}]
    transcriptome = {
        "target_gene": "AAAA" + query + "CCCC",
        "danger_gene": "UUUU" + query + "GGGG",
    }

    results = _score_sirna_batch({
        "batch": batch,
        "target_seq": query,
        "target_seq_id": "target_gene",
        "enabled_rules": ["consensus"],
        "off_target_config": {"level_1_16bp": False, "level_2_20bp": False, "level_3_27bp": False, "seed_7nt": False},
        "transcriptome_seqs": transcriptome,
    })

    assert results[0]["off_target"]["risk_level"] == "low", results[0]
    assert results[0]["off_target"]["risk_score"] == 0, results[0]


def test_phase14_extra_background_uses_distinct_risk_cache(tmp_path):
    cache_dir = tmp_path / "cache"
    query = "AUGCAUGCAUGCAUGCAUGCA"
    primary = TranscriptomeIndex(cache_dir=cache_dir)
    primary.sequences = {"target_gene": "GGGG" + query + "CCCC"}
    primary.source_hash = "primaryhash"
    primary._compute_stats()

    OffTargetRiskIndex.from_transcriptome(primary, cache_dir=cache_dir)
    background = TranscriptomeIndex(cache_dir=cache_dir)
    background.sequences = {"danger_gene": "AAAA" + query + "UUUU"}
    background.source_hash = "backgroundhash"
    background._compute_stats()

    merged = merge_background_transcriptomes(primary, [("host", background)])
    risk = OffTargetRiskIndex.from_transcriptome(merged, cache_dir=cache_dir).assess_sequence(
        query,
        exclude_ids={"target_gene"},
    )

    assert merged.source_hash != primary.source_hash, merged.source_hash
    assert risk["risk_level"] == "high", risk
    assert risk["top_targets"][0]["target_id"] == "host|danger_gene", risk


def test_parallel_long_batch_includes_offtarget_risk():
    batch = [{"sequence": "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC", "start": 0, "end": 50}]
    results = _score_long_dsRNA_batch({
        "batch": batch,
        "enabled_rules": ["consensus"],
        "off_target_config": {"level_1_16bp": True, "level_2_20bp": True, "seed_7nt": True},
        "transcriptome_seqs": {
            "target_gene": batch[0]["sequence"],
            "danger_gene": "GGGG" + batch[0]["sequence"][:25] + "AAAA",
        },
        "target_seq_id": "target_gene",
        "risk_index": None,
    })

    assert "off_target" in results[0], results[0]
    assert results[0]["off_target"]["risk_score"] > 0, results[0]
    assert results[0]["off_target"]["top_targets"], results[0]


def test_exporter_includes_risk_and_validation_columns(tmp_path):
    results = [{
        "rank": 1,
        "sequence": "AUGC",
        "position": "0-4",
        "consensus_score": 80,
        "recommendation_score": 65,
        "cluster_id": 1,
        "cluster_size": 4,
        "alternative_count": 3,
        "cluster_span": "0-7",
        "passed": True,
        "off_target": {
            "risk_level": "medium",
            "risk_score": 30,
            "top_targets": [{"target_id": "gene_x", "risk_score": 30}],
            "validation_direction": "优先验证 gene_x",
        },
    }]
    csv_path = tmp_path / "results.csv"

    ResultExporter().export_csv(results, str(csv_path))
    text = csv_path.read_text(encoding="utf-8")

    assert "off_target_risk" in text
    assert "risk_score" in text
    assert "top_risk_targets" in text
    assert "validation_direction" in text
    assert "cluster_size" in text


def test_phase22_dsrna_exports_explicit_potential_offtarget_genes(tmp_path):
    results = [{
        "rank": 1,
        "sequence": "AUGCAUGCAUGCAUGCAUGCAUGCAUGC",
        "position": "10-38",
        "consensus_score": 82,
        "recommendation_score": 42,
        "passed": True,
        "off_target": {
            "risk_level": "high",
            "risk_score": 95,
            "top_targets": [
                {"target_id": "danger_gene", "risk_score": 95, "reasons": ["20bp continuous match"]},
                {"target_id": "seed_gene", "risk_score": 35, "reasons": ["7nt seed match"]},
            ],
            "validation_direction": "优先验证 danger_gene；备选验证 seed_gene",
        },
        "validation_hits": [
            {
                "target_id": "danger_gene",
                "risk_score": 95,
                "match_type": "20bp_contiguous",
                "target_position": 120,
                "longest_contiguous_match": 20,
                "mismatch_count": 1,
                "seed_match": True,
                "validation_action": "High priority: avoid or validate by BLAST/qPCR",
                "query_fragment": "AUGCAUGCAUGCAUGCAUGC",
                "target_fragment": "AUGCAUGCAUGCAUGCAUGC",
            }
        ],
    }]
    csv_path = tmp_path / "results.csv"
    report_path = tmp_path / "validation_report.xlsx"

    exporter = ResultExporter()
    exporter.export_csv(results, str(csv_path))
    exporter.export_validation_report(results, str(report_path))

    csv_text = csv_path.read_text(encoding="utf-8")
    assert "top_offtarget_genes" in csv_text
    assert "danger_gene" in csv_text and "20bp continuous match" in csv_text

    import openpyxl

    workbook = openpyxl.load_workbook(report_path)
    recommendations = list(workbook["Recommendations"].iter_rows(values_only=True))
    header = recommendations[0]
    row = recommendations[1]
    assert "Top Off-target Genes" in header
    assert "danger_gene" in row[header.index("Top Off-target Genes")]


def test_parallel_progress_keeps_room_for_ranking_and_saving():
    index = TranscriptomeIndex()
    seq = "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
    index.sequences = {"target_gene": seq, "background": "UUUUUUUUUUUUUUUUUUUUUUUUUUUUUU"}
    index._compute_stats()
    messages = []
    config = DesignConfig(mode="long_dsRNA", length_min=21, length_max=30, n_cores=2, batch_size=5)

    task = ParallelDesignTask(db_manager=None, n_cores=2, batch_size=5)
    task.run_parallel(index, "target_gene", config, progress_callback=lambda step, pct: messages.append((step, pct)))

    batch_percents = [pct for step, pct in messages if step.startswith("Batch ")]
    assert batch_percents and max(batch_percents) < 90, messages
    assert any("Ranking" in step for step, _ in messages), messages
    assert any("Saving" in step for step, _ in messages), messages


def test_phase18_parallel_task_persists_full_design_config():
    import json

    index = TranscriptomeIndex()
    seq = "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
    index.sequences = {"target_gene": seq, "background": "UUUUUUUUUUUUUUUUUUUUUUUUUUUUUU"}
    index._compute_stats()
    db = DatabaseManager(":memory:")
    config = DesignConfig(
        mode="siRNA",
        gc_min=25,
        gc_max=75,
        enabled_rules=["consensus", "reynolds"],
        n_cores=2,
        batch_size=5,
        preset="relaxed",
    )

    result = ParallelDesignTask(db_manager=db, n_cores=2, batch_size=5).run_parallel(
        index,
        "target_gene",
        config,
    )

    task = db.get_task(result["task_id"])
    params = json.loads(task["params_json"])
    assert params["mode"] == "siRNA", params
    assert params["gc_min"] == 25, params
    assert params["gc_max"] == 75, params
    assert params["enabled_rules"] == ["consensus", "reynolds"], params
    assert params["preset"] == "relaxed", params


def test_phase19_dsirna_keeps_continuous_offtarget_when_seed_disabled():
    query = "AUGCAUGCAUGCAUGCAUGCAUGCAUG"
    index = TranscriptomeIndex()
    index.sequences = {
        "target_gene": "AAAA" + query + "CCCC",
        "danger_gene": "GGGG" + query[:21] + "UUUU",
    }
    index._compute_stats()
    config = DesignConfig(
        mode="DsiRNA",
        gc_min=0,
        gc_max=100,
        exclude_poly_n=9999,
        off_target_levels={
            "level_1_16bp": True,
            "level_2_20bp": True,
            "level_3_27bp": False,
            "seed_7nt": False,
        },
    )

    result = DesignTask(DatabaseManager(":memory:")).run(index, "target_gene", config)

    assert result["results"], result
    risky = [
        item for item in result["results"]
        if (item.get("off_target") or {}).get("risk_score", 0) > 0
    ]
    assert risky, result["results"]
    assert any(
        target.get("target_id") == "danger_gene"
        for item in risky
        for target in (item.get("off_target") or {}).get("top_targets", [])
    ), result["results"]


def test_rnaup_records_explicit_fallback_when_cli_is_absent():
    calc = ThermodynamicsCalculator.__new__(ThermodynamicsCalculator)
    calc.available = True
    calc.rnaup_executable = None
    calc.rnaduplex = lambda seq1, seq2: {"dg": -9.1, "structure": "(((&)))"}

    result = calc.rnaup("AUGCAUGC", "GCAUGCAU")

    assert result["dg"] == -9.1
    assert result["details"]["method"] == "RNAduplex-fallback-for-RNAup"
    assert result["details"]["rnaup_available"] is False


def test_redundancy_clusters_one_bp_sliding_neighbors():
    results = [
        {
            "rank": 1,
            "sequence": "A" * 21,
            "position_start": 0,
            "position_end": 21,
            "consensus_score": 80,
            "recommendation_score": 80,
        },
        {
            "rank": 2,
            "sequence": "C" * 21,
            "position_start": 1,
            "position_end": 22,
            "consensus_score": 79,
            "recommendation_score": 79,
        },
        {
            "rank": 3,
            "sequence": "G" * 21,
            "position_start": 50,
            "position_end": 71,
            "consensus_score": 70,
            "recommendation_score": 70,
        },
    ]

    clustered = cluster_redundant_results(results, overlap_threshold=0.8)

    assert len(clustered) == 2, clustered
    assert clustered[0]["cluster_size"] == 2, clustered
    assert clustered[0]["alternative_count"] == 1, clustered
    assert clustered[1]["cluster_size"] == 1, clustered


def test_design_summary_reports_raw_and_nonredundant_counts():
    seq = "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": seq, "background": "UUUUUUUUUUUUUUUUUUUUUUUUUUUU"}
    index._compute_stats()

    config = DesignConfig(mode="siRNA", n_cores=1)
    result = DesignTask().run(index, "target_gene", config)

    summary = result["summary"]
    assert summary["raw_candidates"] > summary["nonredundant_candidates"], summary
    assert all(r.get("cluster_size", 0) >= 1 for r in result["results"]), result["results"][:3]


def test_phase12_result_explanation_validation_hits_and_region_map():
    query = "AUGCAUGCAUGCAUGCAUGCA"
    index = TranscriptomeIndex()
    index.sequences = {
        "target_gene": "GGGG" + query + "CCCC" + "U" * 40,
        "danger_gene": "AAAA" + query + "UUUU" + "G" * 40,
    }
    index._compute_stats()
    off_target = OffTargetScreener(index).screen_sequence(
        query,
        exclude_ids={"target_gene"},
        use_vienna=False,
    )
    result = {
        "rank": 1,
        "sequence": query,
        "position": "4-25",
        "position_start": 4,
        "position_end": 25,
        "consensus_score": 82,
        "recommendation_score": 42,
        "cluster_size": 3,
        "passed": False,
        "off_target": off_target,
        "rnaup": {"dg": -9.2, "details": {"method": "RNAduplex-fallback-for-RNAup"}},
    }

    explanation = explain_result(result, mode="siRNA")
    assert "推荐理由" in explanation["summary"]
    assert any("danger_gene" in item for item in explanation["risk_notes"]), explanation
    assert any("RNAduplex fallback" in item for item in explanation["method_notes"]), explanation

    hits = build_validation_hits(result, index, max_hits=3)
    assert hits[0]["target_id"] == "danger_gene", hits
    assert hits[0]["longest_contiguous_match"] >= 20, hits
    assert hits[0]["mismatch_count"] == 0, hits
    assert hits[0]["query_fragment"] == query, hits

    region_map = render_region_map(target_length=80, start=4, end=25, risk_positions=[4])
    assert "candidate" in region_map
    assert "#" in region_map
    assert "!" in region_map


def test_phase12_t7_primer_design_for_long_dsrna():
    sequence = (
        "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
        "GGCCAAUUGGCCAAUUGGCCAAUUGGCCAAUUGGCCAAUU"
        "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
        "GGCCAAUUGGCCAAUUGGCCAAUUGGCCAAUUGGCCAAUU"
        "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC"
    )

    primer_set = design_t7_primers(sequence, result_id="candidate_1")

    assert primer_set["product_length"] == len(sequence)
    assert primer_set["forward_primer"]["sequence"]
    assert primer_set["reverse_primer"]["sequence"]
    assert primer_set["t7_forward_primer"]["sequence"].startswith(T7_PROMOTER)
    assert primer_set["t7_reverse_primer"]["sequence"].startswith(T7_PROMOTER)
    assert 45 <= primer_set["forward_primer"]["tm"] <= 75
    assert 20 <= primer_set["forward_primer"]["gc_percent"] <= 70


def test_phase17_primer_specificity_checks_reverse_complement_strand_and_three_prime_anchor():
    primer = "ATGCGTACGTAGCTAGCTAG"
    reverse_binding_site = "CTAGCTAGCTACGTACGCAT"
    three_prime_risk_site = "TTTCGTACGTAGCTAGCTAG"
    three_prime_mismatch_site = "ATGCGTACGTAGCTAGCTAA"
    transcriptome = {
        "reverse_offtarget": "AAAA" + reverse_binding_site + "UUUU",
        "three_prime_risk": "GGGG" + three_prime_risk_site + "CCCC",
        "full_near_match": "CCCC" + three_prime_mismatch_site + "AAAA",
    }

    result = check_primer_specificity(primer, transcriptome)

    assert result["specific"] is False, result
    assert any(hit["target_id"] == "reverse_offtarget" and hit["strand"] == "-" for hit in result["matches"]), result
    assert any(hit["target_id"] == "three_prime_risk" and hit["three_prime_match_len"] >= 10 for hit in result["matches"]), result
    assert any(hit["target_id"] == "full_near_match" and hit["mismatches"] == 1 for hit in result["matches"]), result


def test_phase18_primer_specificity_keeps_duplicate_anchor_offsets():
    primer = "AAAAAAACCCCCCAAAAAAA"
    near_match = "TAAAAAACCCCCGAAAAAAA"
    transcriptome = {"duplicate_anchor_near_match": "GGGG" + near_match + "UUUU"}

    result = check_primer_specificity(primer, transcriptome)

    assert any(
        hit["target_id"] == "duplicate_anchor_near_match" and hit["mismatches"] == 2
        for hit in result["matches"]
    ), result


def test_phase17_primer_tm_uses_nearest_neighbor_not_wallace():
    primer = "ATGCGTACGTAGCTAGCTAG"
    wallace = 2 * (primer.count("A") + primer.count("T")) + 4 * (primer.count("G") + primer.count("C"))

    tm = calculate_primer_tm(primer)

    assert 45 <= tm <= 75, tm
    assert abs(tm - wallace) > 1.0, (tm, wallace)


def test_phase17_offtarget_thermo_does_not_silently_truncate_by_fasta_order():
    query = "AUGCAUGCAUGCAUGCAUGCA"
    seed = query[1:8]
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": query}
    for i in range(5000):
        index.sequences[f"filler_{i:04d}"] = "U" * 40
    index.sequences["late_seed_hit"] = "AAAA" + seed + "UUUU"
    index._compute_stats()

    class FakeThermo:
        available = True

        def rnaduplex(self, seq1, seq2):
            return {"dg": -8.2, "structure": "fake"}

    screener = OffTargetScreener(index)
    screener._thermo = FakeThermo()
    result = screener.screen_sequence(query, exclude_ids={"target_gene"}, use_vienna=True)

    thermo = result["thermodynamics"]
    assert not thermo.get("truncated"), thermo
    assert any(hit["target_id"] == "late_seed_hit" for hit in thermo["seed_hits"]), thermo


def test_phase17_candidate_budget_counts_only_clean_windows():
    task = DesignTask(DatabaseManager(":memory:"))
    assert task._candidate_window_count("N" * 100, 21, 21) == 0
    assert task._candidate_window_count("AUGCAUGCAUGCAUGCAUGCA", 21, 21) == 1


def test_phase12_three_presets_adjust_config_without_exposing_rules():
    strict = apply_preset_to_config(DesignConfig(mode="long_dsRNA"), "strict")
    balanced = apply_preset_to_config(DesignConfig(mode="long_dsRNA"), "balanced")
    relaxed = apply_preset_to_config(DesignConfig(mode="long_dsRNA"), "relaxed")

    assert strict.preset == "strict"
    assert balanced.preset == "balanced"
    assert relaxed.preset == "relaxed"
    assert strict.gc_min >= balanced.gc_min >= relaxed.gc_min
    assert strict.gc_max <= balanced.gc_max <= relaxed.gc_max
    assert strict.thermodynamics["rnaup_top_n"] >= balanced.thermodynamics["rnaup_top_n"]
    assert relaxed.off_target_levels["level_3_27bp"] is False


def test_phase12_project_file_roundtrip(tmp_path):
    project_path = tmp_path / "experiment.dsforge_project"
    config = DesignConfig(mode="siRNA", n_cores=2)
    payload = {
        "transcriptome": {"source_file": "species.fa", "cache_key": "abc"},
        "target": {"source": "transcriptome", "id": "gene_a"},
        "config": config,
        "results": [{"rank": 1, "sequence": "AUGC", "position": "0-4"}],
    }

    save_project_file(str(project_path), payload)
    loaded = load_project_file(str(project_path))

    assert loaded["version"] == 1
    assert loaded["config"]["mode"] == "siRNA"
    assert loaded["config"]["n_cores"] == 2
    assert loaded["results"][0]["sequence"] == "AUGC"


def test_phase12_multi_background_merge_prefixes_ids():
    primary = TranscriptomeIndex()
    primary.sequences = {"target_gene": "AUGCAUGCAUGCAUGCAUGC"}
    primary._compute_stats()
    host = TranscriptomeIndex()
    host.sequences = {"actin": "UUUUUUUUUUUUUUUUUUUU"}
    host._compute_stats()

    merged = merge_background_transcriptomes(primary, [("host plant", host)])

    assert "target_gene" in merged.sequences
    assert "host_plant|actin" in merged.sequences
    assert merged.get_stats()["num_sequences"] == 2


def test_phase12_cache_manifest_can_rename_and_delete_saved(tmp_path):
    cache_dir = tmp_path / "cache"
    fasta = tmp_path / "species.fa"
    fasta.write_text(">gene_a\nAUGCAUGCAUGCAUGCAUGC\n", encoding="utf-8")

    index = TranscriptomeIndex(cache_dir=cache_dir).load_fasta(str(fasta))
    key = index.source_hash
    TranscriptomeIndex.rename_saved(key, "Renamed species", cache_dir=cache_dir)
    saved = TranscriptomeIndex.list_saved(cache_dir=cache_dir)
    assert saved[0]["name"] == "Renamed species"

    TranscriptomeIndex.delete_saved(key, cache_dir=cache_dir, delete_cache=True)
    assert TranscriptomeIndex.list_saved(cache_dir=cache_dir) == []
    assert not Path(index.cache_path).exists()


def test_phase12_validation_report_export_includes_sheets(tmp_path):
    results = [{
        "rank": 1,
        "sequence": "AUGCAUGCAUGCAUGCAUGCA",
        "position": "0-21",
        "consensus_score": 81,
        "recommendation_score": 80,
        "passed": True,
        "explanation": {"summary": "推荐理由：综合分较高"},
        "validation_hits": [{"target_id": "gene_x", "match_type": "seed_7nt", "validation_action": "qPCR"}],
        "primers": {"product_length": 21, "forward_primer": {"sequence": "ATGC"}, "reverse_primer": {"sequence": "GCAT"}},
        "off_target": {"risk_level": "low", "risk_score": 0, "top_targets": [], "validation_direction": "常规验证"},
    }]
    report_path = tmp_path / "validation_report.xlsx"

    ResultExporter().export_validation_report(results, str(report_path))

    import openpyxl

    workbook = openpyxl.load_workbook(report_path)
    assert {"Recommendations", "OffTarget Validation", "Primers", "Methods"} <= set(workbook.sheetnames)


def test_phase13_sgrna_scans_spcas9_ngg_on_both_strands():
    seq = "A" * 25 + "G" * 20 + "AGG" + "T" * 12 + "CCG" + "C" * 20 + "A" * 25

    candidates = scan_sgrna_candidates(seq)

    assert any(c["pam"] == "AGG" and c["strand"] == "+" for c in candidates), candidates
    reverse_hits = [c for c in candidates if c["strand"] == "-"]
    assert any(c["pam"].endswith("GG") and c["genomic_pam"] == "CCG" for c in reverse_hits), candidates
    assert all(len(c["spacer_dna"]) == 20 for c in candidates), candidates
    assert all(c["pam"][-2:] == "GG" for c in candidates), candidates


def test_phase16_reverse_sgrna_cut_site_uses_pam_proximal_spcas9_coordinate():
    seq = "A" * 10 + "CCG" + "C" * 20 + "A" * 10

    reverse_hit = next(c for c in scan_sgrna_candidates(seq) if c["strand"] == "-" and c["genomic_pam"] == "CCG")

    assert reverse_hit["pam"] == "CGG", reverse_hit
    assert reverse_hit["pam_start"] == 10, reverse_hit
    assert reverse_hit["position_start"] == 13, reverse_hit
    assert reverse_hit["position_end"] == 33, reverse_hit
    assert reverse_hit["cut_site"] == 16, reverse_hit


def test_phase13_sgrna_offtarget_ranking_excludes_intended_site():
    spacer = "G" * 20
    target = "A" * 25 + spacer + "AGG" + "T" * 40
    near = "C" * 20 + "G" * 19 + "A" + "TGG" + "C" * 20
    reference = {
        "target_gene": target,
        "near_offtarget": near,
    }
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)

    risk = score_sgrna_offtargets(candidate, reference, exclude_target_id="target_gene")

    assert risk["risk_level"] in {"medium", "high"}, risk
    assert risk["top_targets"][0]["target_id"] == "near_offtarget", risk
    assert risk["top_targets"][0]["mismatches"] == 1, risk
    assert risk["top_targets"][0]["cfd_like_score"] > 0, risk


def test_phase19_sgrna_offtarget_checks_other_sites_in_same_reference_sequence():
    spacer = "G" * 20
    near_same_sequence = "G" * 19 + "A"
    target = "A" * 25 + spacer + "AGG" + "T" * 20 + near_same_sequence + "TGG" + "C" * 20
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)

    risk = score_sgrna_offtargets(candidate, {"target_gene": target}, exclude_target_id="target_gene")

    assert risk["risk_level"] in {"medium", "high"}, risk
    assert risk["top_targets"][0]["target_id"] == "target_gene", risk
    assert risk["top_targets"][0]["mismatches"] == 1, risk
    assert risk["top_targets"][0]["position"] != candidate["position_start"], risk


def test_phase20_sgrna_offtarget_counts_nrg_and_seed_pot_hits():
    seed = "GGGGTTTTAAAA"
    spacer = "AAAACCCC" + seed
    seed_matched_offtarget = "TTTTCCCC" + seed
    target = "A" * 25 + spacer + "AGG" + "C" * 30
    reference = {
        "target_gene": target,
        "nrg_seed_hit": "C" * 20 + seed_matched_offtarget + "AAG" + "G" * 20,
    }
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)

    risk = score_sgrna_offtargets(candidate, reference, exclude_target_id="target_gene")

    assert risk["summary"]["mismatch_counts"]["4M"] == 1, risk
    assert risk["summary"]["pot_mismatch_counts"]["4M"] == 1, risk
    assert risk["summary"]["nrg_pam_hits"] == 1, risk
    assert risk["summary"]["nag_pam_hits"] == 1, risk
    assert risk["top_targets"][0]["pam"] == "AAG", risk
    assert risk["top_targets"][0]["seed_12_match"] is True, risk
    assert any(match["match_type"] == "Cas9_seed12_POT_4mm" for match in risk["matches"]), risk


def test_phase20_sgrna_offtarget_detects_reverse_strand_nag_sites():
    spacer = "ACGTACGTACGTACGTACGT"
    reverse_offtarget_spacer = "ACGTACGTACGTACGTACGA"
    target = "A" * 25 + spacer + "AGG" + "T" * 20
    genomic_pam = "CTT"  # reverse complement of guide-oriented AAG
    reverse_site = genomic_pam + reverse_offtarget_spacer.translate(str.maketrans("ACGT", "TGCA"))[::-1]
    reference = {
        "target_gene": target,
        "reverse_nag": "C" * 10 + reverse_site + "G" * 20,
    }
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)

    risk = score_sgrna_offtargets(candidate, reference, exclude_target_id="target_gene")

    reverse_hits = [
        hit for hit in risk["top_targets"]
        if hit["target_id"] == "reverse_nag" and hit["strand"] == "-" and hit["pam"] == "AAG"
    ]
    assert reverse_hits, risk
    assert reverse_hits[0]["genomic_pam"] == genomic_pam, risk
    assert risk["summary"]["nag_pam_hits"] >= 1, risk


def test_phase20_sgrna_risk_evaluation_uses_offtarget_not_on_target_zero_mismatch_semantics():
    spacer = "G" * 20
    target = "A" * 25 + spacer + "AGG" + "T" * 40
    near = "C" * 20 + "G" * 19 + "A" + "TGG" + "C" * 20
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)

    risk = score_sgrna_offtargets(
        candidate,
        {"target_gene": target, "near_offtarget": near},
        exclude_target_id="target_gene",
    )

    assert risk["summary"]["mismatch_counts"]["0M"] == 0, risk
    assert risk["summary"]["mismatch_counts"]["1M"] >= 1, risk
    assert risk["summary"]["sgrnacas9_risk_evaluation"] == "High_risk", risk


def test_phase20_validation_hits_preserve_zero_based_window_and_use_protospacer_mismatches():
    spacer = "G" * 20
    target = "A" * 25 + spacer + "AGG" + "T" * 40
    near = "C" * 20 + "G" * 19 + "A" + "TGG" + "C" * 20
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": target, "near_offtarget": near}
    index._compute_stats()
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)
    risk = score_sgrna_offtargets(candidate, index.sequences, exclude_target_id="target_gene")

    hits = build_validation_hits({"sequence": candidate["guide_rna"], "off_target": risk}, index, max_hits=1)

    assert hits[0]["validation_window_start"] == 0, hits
    assert hits[0]["target_fragment"].startswith("C" * 20), hits
    assert hits[0]["target_protospacer_pam"].endswith("TGG"), hits
    assert hits[0]["mismatch_count"] == 1, hits


def test_phase20_validation_hits_keep_distinct_sgrna_loci_in_same_target_and_prefer_pot_label():
    seed = "GGGGTTTTAAAA"
    spacer = "AAAACCCC" + seed
    first = "TTTTCCCC" + seed
    second = "AAAACCCC" + seed[:-1] + "T"
    target = "A" * 25 + spacer + "AGG" + "C" * 30
    repeated_offtarget = "N" * 0 + "C" * 20 + first + "AAG" + "G" * 30 + second + "TGG" + "C" * 20
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": target, "repeat_target": repeated_offtarget}
    index._compute_stats()
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)
    risk = score_sgrna_offtargets(candidate, index.sequences, exclude_target_id="target_gene")

    hits = [
        hit for hit in build_validation_hits({"sequence": candidate["guide_rna"], "off_target": risk}, index, max_hits=5)
        if hit["target_id"] == "repeat_target"
    ]

    assert len({hit["target_position"] for hit in hits}) >= 2, hits
    assert any(hit["match_type"].startswith("Cas9_seed12_POT") for hit in hits), hits


def test_phase20_sgrna_cloning_oligos_warn_for_vector_restriction_sites_and_custom_overhangs():
    spacer = "GGGTCTCAAAACCCCGGGTT"

    oligos = design_sgrna_cloning_oligos(spacer, forward_overhang="ACCG", reverse_overhang="AAAC")

    assert oligos["custom_forward_oligo"]["sequence"].startswith("ACCG"), oligos
    assert oligos["custom_reverse_oligo"]["sequence"].startswith("AAAC"), oligos
    assert any(warning["enzyme"] == "BsaI" for warning in oligos["restriction_site_warnings"]), oligos
    assert "restriction enzyme" in oligos["notes"], oligos


def test_phase20_sgrna_export_includes_sgrnacas9_style_ot_and_pot_counts(tmp_path):
    risk = {
        "risk_level": "medium",
        "risk_score": 48,
        "top_targets": [{"target_id": "gene_x", "risk_score": 48}],
        "validation_direction": "对 Top Cas9 off-target 做扩增测序。",
        "summary": {
            "mismatch_counts": {"0M": 1, "1M": 0, "2M": 1, "3M": 0, "4M": 2, "5M": 3},
            "pot_mismatch_counts": {"0M": 1, "1M": 0, "2M": 0, "3M": 0, "4M": 1, "5M": 0},
            "total_ot": 6,
            "total_pot": 1,
            "nrg_pam_hits": 2,
            "nag_pam_hits": 2,
        },
    }
    csv_path = tmp_path / "sgrna_counts.csv"

    ResultExporter().export_csv([{
        "rank": 1,
        "sequence": "G" * 20,
        "position": "25-45",
        "consensus_score": 80,
        "passed": True,
        "off_target": risk,
        "sgrna": {"spacer_dna": "G" * 20, "guide_rna": "G" * 20, "pam": "AGG"},
    }], str(csv_path))
    text = csv_path.read_text(encoding="utf-8")

    assert "sgrna_ot_0M" in text and "sgrna_ot_5M" in text, text
    assert "sgrna_pot_0M" in text and "sgrna_pot_5M" in text, text
    assert "sgrna_total_pot" in text and "sgrna_nrg_pam_hits" in text, text
    assert "sgrna_nag_pam_hits" in text, text


def test_phase20_sgrna_design_task_collapses_adjacent_overlapping_guides():
    seq = "A" * 25 + "G" * 26 + "AGG" + "T" * 40
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": seq, "background": "C" * 120}
    index._compute_stats()

    result = DesignTask(DatabaseManager(":memory:")).run(index, "target_gene", DesignConfig(mode="sgRNA", n_cores=1))

    assert result["summary"]["raw_candidates"] > result["summary"]["nonredundant_candidates"], result["summary"]
    assert any(item.get("cluster_size", 1) > 1 for item in result["results"]), result["results"]


def test_phase13_sgrna_cloning_oligos_and_report_export(tmp_path):
    oligos = design_sgrna_cloning_oligos("G" * 20)

    assert oligos["u6_spacer_dna"] == "G" * 20
    assert oligos["px330_forward_oligo"]["sequence"].startswith("CACCG")
    assert oligos["px330_reverse_oligo"]["sequence"].startswith("AAAC")
    assert oligos["guide_rna"] == "G" * 20

    results = [{
        "rank": 1,
        "sequence": "G" * 20,
        "position": "25-45",
        "consensus_score": 88,
        "recommendation_score": 86,
        "passed": True,
        "sgrna": {
            "spacer_dna": "G" * 20,
            "pam": "AGG",
            "strand": "+",
            "cut_site": 42,
            "cloning_oligos": oligos,
        },
        "off_target": {"risk_level": "low", "risk_score": 0, "top_targets": [], "validation_direction": "常规验证"},
        "validation_hits": [],
        "primers": {"sgrna_cloning_oligos": oligos},
    }]
    report_path = tmp_path / "sgrna_report.xlsx"
    ResultExporter().export_validation_report(results, str(report_path))

    import openpyxl

    workbook = openpyxl.load_workbook(report_path)
    primer_rows = list(workbook["Primers"].iter_rows(values_only=True))
    assert any("CACCG" in str(cell) for row in primer_rows for cell in row if cell), primer_rows


def test_phase15_sgrna_export_and_history_preserve_pam_strand_cut_site(tmp_path):
    seq = "A" * 25 + "G" * 20 + "AGG" + "T" * 40
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": seq, "background": "C" * 100}
    index._compute_stats()

    result = DesignTask(DatabaseManager(":memory:")).run(index, "target_gene", DesignConfig(mode="sgRNA", n_cores=1))
    top = result["results"][0]
    csv_path = tmp_path / "sgrna_results.csv"
    ResultExporter().export_csv([top], str(csv_path))
    csv_text = csv_path.read_text(encoding="utf-8")

    loaded = DesignTask(DatabaseManager(":memory:")).db
    task_id = loaded.create_task("sgRNA", "target_gene", seq, {})
    DesignTask(loaded)._save_results(task_id, [top], DesignConfig(mode="sgRNA"))
    history_top = loaded.get_results(task_id)[0]

    assert "sgrna_pam" in csv_text, csv_text
    assert "sgrna_strand" in csv_text, csv_text
    assert "sgrna_cut_site" in csv_text, csv_text
    assert history_top["sgrna"]["pam"].endswith("GG"), history_top
    assert isinstance(history_top["sgrna"]["cut_site"], int), history_top


def test_phase13_design_task_runs_sgrna_mode():
    seq = "A" * 25 + "G" * 20 + "AGG" + "T" * 40
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": seq, "background": "C" * 100}
    index._compute_stats()

    result = DesignTask().run(index, "target_gene", DesignConfig(mode="sgRNA", n_cores=1))

    assert result["summary"]["total_candidates"] >= 1, result
    top = result["results"][0]
    assert "sgrna" in top, top
    assert top["sgrna"]["pam"].endswith("GG"), top
    assert "cloning_oligos" in top["sgrna"], top


def test_phase14_history_reload_preserves_rnaup_method():
    db = DatabaseManager(":memory:")
    task_id = db.create_task(
        mode="siRNA",
        target_seq_id="target_gene",
        target_seq="AUGCAUGCAUGCAUGCAUGC",
        params={},
    )
    DesignTask(db)._save_results(task_id, [{
        "rank": 1,
        "sequence": "AUGCAUGCAUGCAUGCAUGC",
        "position_start": 0,
        "position_end": 21,
        "consensus_score": 80,
        "passed": True,
        "off_target": {"risk_level": "low", "risk_score": 0, "top_targets": [], "validation_direction": "常规验证"},
        "rnaup": {"dg": -8.1, "details": {"method": "RNAup-cli"}},
    }], DesignConfig(mode="siRNA"))

    loaded = db.get_results(task_id)[0]

    assert loaded["rnaup"]["details"]["method"] == "RNAup-cli", loaded
    assert loaded["rnaup"]["dg"] == -8.1, loaded


def test_phase16_database_delete_task_cascades_child_rows(tmp_path):
    db = DatabaseManager(str(tmp_path / "cascade.db"))
    task_id = db.create_task("siRNA", "target_gene", "AUGCAUGCAUGCAUGCAUGC", {})
    result_id = db.add_result(task_id, 1, "AUGCAUGCAUGCAUGCAUGC", 0, 21, 80, 1)
    db.add_rule_score(result_id, "consensus", 80, True, [])
    db.add_thermodynamics(result_id, on_target_dg=-10.0)
    db.add_pool_detail(result_id, "AUGCAUGCAUGCAUGCAUGC", 21, 80)

    db.delete_task(task_id)

    assert db.get_results(task_id) == []
    assert db.get_rule_scores(result_id) == []
    with db._get_conn() as conn:
        thermal_count = conn.execute("SELECT COUNT(*) FROM thermodynamics WHERE result_id = ?", (result_id,)).fetchone()[0]
        pool_count = conn.execute("SELECT COUNT(*) FROM pool_details WHERE result_id = ?", (result_id,)).fetchone()[0]
    assert thermal_count == 0
    assert pool_count == 0


def test_phase19_schema_upgrade_preserves_results_foreign_key_for_old_databases():
    import sqlite3

    conn = sqlite3.connect(":memory:")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(
        """
        CREATE TABLE design_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mode TEXT NOT NULL CHECK(mode IN ('siRNA', 'DsiRNA', 'long_dsRNA')),
            target_seq_id TEXT,
            target_seq TEXT,
            params_json TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN ('pending', 'running', 'completed', 'cancelled', 'failed')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP
        );
        CREATE TABLE results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            rank INTEGER NOT NULL,
            candidate_seq TEXT NOT NULL,
            position_start INTEGER,
            position_end INTEGER,
            consensus_score REAL,
            passed_filters INTEGER DEFAULT 0,
            FOREIGN KEY (task_id) REFERENCES design_tasks(id) ON DELETE CASCADE
        );
        """
    )

    init_database(conn)

    fk_targets = [row[2] for row in conn.execute("PRAGMA foreign_key_list(results)").fetchall()]
    assert fk_targets == ["design_tasks"], fk_targets
    conn.execute(
        "INSERT INTO design_tasks (mode, target_seq_id, target_seq, params_json, status) VALUES (?, ?, ?, ?, ?)",
        ("sgRNA", "target_gene", "A" * 40, "{}", "completed"),
    )
    task_id = conn.execute("SELECT id FROM design_tasks").fetchone()[0]
    conn.execute(
        "INSERT INTO results (task_id, rank, candidate_seq, position_start, position_end, consensus_score) VALUES (?, ?, ?, ?, ?, ?)",
        (task_id, 1, "G" * 20, 0, 20, 80),
    )
    conn.execute("DELETE FROM design_tasks WHERE id = ?", (task_id,))
    conn.commit()

    assert conn.execute("SELECT COUNT(*) FROM results").fetchone()[0] == 0


def test_phase16_design_cancellation_marks_task_cancelled():
    index = TranscriptomeIndex()
    index.sequences = {
        "target_gene": "AUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGCAUGC",
        "background": "U" * 80,
    }
    index._compute_stats()
    db = DatabaseManager(":memory:")

    def cancel_on_first_progress(_step, _percent):
        raise DesignCancelled("test cancellation")

    try:
        DesignTask(db).run(index, "target_gene", DesignConfig(mode="siRNA"), progress_callback=cancel_on_first_progress)
        assert False, "DesignCancelled was not raised"
    except DesignCancelled:
        pass

    assert db.list_tasks(1)[0]["status"] == "cancelled"


def test_phase16_long_dsrna_large_target_is_rejected_before_materializing_candidates():
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": "AUGC" * 2500}
    index._compute_stats()
    config = DesignConfig(mode="long_dsRNA", length_min=200, length_max=500, max_raw_candidates=1000)

    try:
        DesignTask(DatabaseManager(":memory:")).run(index, "target_gene", config)
        assert False, "Expected candidate limit ValueError"
    except ValueError as exc:
        assert "Too many candidate windows" in str(exc), exc


def test_phase13_sgrna_validation_hits_use_amplicon_sequencing_action():
    spacer = "G" * 20
    target = "A" * 25 + spacer + "AGG" + "T" * 40
    near = "C" * 20 + "G" * 19 + "A" + "TGG" + "C" * 20
    index = TranscriptomeIndex()
    index.sequences = {"target_gene": target, "near_offtarget": near}
    index._compute_stats()
    candidate = next(c for c in scan_sgrna_candidates(target) if c["spacer_dna"] == spacer)
    off_target = score_sgrna_offtargets(candidate, index.sequences, exclude_target_id="target_gene")
    result = {
        "sequence": spacer.replace("T", "U"),
        "off_target": off_target,
    }

    hits = build_validation_hits(result, index, max_hits=3)

    assert hits, off_target
    assert hits[0]["match_type"].startswith("Cas9_"), hits
    assert "ICE/TIDE" in hits[0]["validation_action"], hits


if __name__ == "__main__":
    test_offtarget_excludes_intended_target_sequence()
    test_parallel_worker_loads_builtin_scoring_rules()
    test_gui_worker_uses_parallel_task_when_multiple_cores_requested()
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as tmpdir:
        test_load_first_target_fasta_record_normalizes_sequence(Path(tmpdir))
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase16_load_fasta_rejects_empty_or_headerless_files(Path(tmpdir))
    test_phase16_generate_candidates_skips_ambiguous_windows_but_keeps_clean_flanks()
    test_phase17_ambiguous_iupac_windows_are_reported_in_diagnosis()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_transcriptome_cache_manifest_loads_saved_without_original_file(Path(tmpdir))
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase16_transcriptome_and_offtarget_caches_are_json_not_pickle(Path(tmpdir))
    test_no_result_diagnosis_explains_short_target()
    test_phase15_sgrna_no_result_diagnosis_explains_missing_spcas9_pam()
    test_clone_with_custom_sequence_keeps_background_and_adds_target()
    test_offtarget_risk_summary_ranks_targets_and_suggests_validation()
    test_phase14_offtarget_rule_switches_are_honored()
    test_phase14_parallel_worker_honors_offtarget_config()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase14_extra_background_uses_distinct_risk_cache(Path(tmpdir))
    test_parallel_long_batch_includes_offtarget_risk()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_exporter_includes_risk_and_validation_columns(Path(tmpdir))
    test_parallel_progress_keeps_room_for_ranking_and_saving()
    test_phase18_parallel_task_persists_full_design_config()
    test_phase19_dsirna_keeps_continuous_offtarget_when_seed_disabled()
    test_rnaup_records_explicit_fallback_when_cli_is_absent()
    test_redundancy_clusters_one_bp_sliding_neighbors()
    test_design_summary_reports_raw_and_nonredundant_counts()
    test_phase12_result_explanation_validation_hits_and_region_map()
    test_phase12_t7_primer_design_for_long_dsrna()
    test_phase17_primer_specificity_checks_reverse_complement_strand_and_three_prime_anchor()
    test_phase18_primer_specificity_keeps_duplicate_anchor_offsets()
    test_phase17_primer_tm_uses_nearest_neighbor_not_wallace()
    test_phase17_offtarget_thermo_does_not_silently_truncate_by_fasta_order()
    test_phase17_candidate_budget_counts_only_clean_windows()
    test_phase12_three_presets_adjust_config_without_exposing_rules()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase12_project_file_roundtrip(Path(tmpdir))
    test_phase12_multi_background_merge_prefixes_ids()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase12_cache_manifest_can_rename_and_delete_saved(Path(tmpdir))
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase12_validation_report_export_includes_sheets(Path(tmpdir))
    test_phase13_sgrna_scans_spcas9_ngg_on_both_strands()
    test_phase16_reverse_sgrna_cut_site_uses_pam_proximal_spcas9_coordinate()
    test_phase13_sgrna_offtarget_ranking_excludes_intended_site()
    test_phase19_sgrna_offtarget_checks_other_sites_in_same_reference_sequence()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase13_sgrna_cloning_oligos_and_report_export(Path(tmpdir))
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase15_sgrna_export_and_history_preserve_pam_strand_cut_site(Path(tmpdir))
    test_phase13_design_task_runs_sgrna_mode()
    test_phase14_history_reload_preserves_rnaup_method()
    with tempfile.TemporaryDirectory() as tmpdir:
        test_phase16_database_delete_task_cascades_child_rows(Path(tmpdir))
    test_phase19_schema_upgrade_preserves_results_foreign_key_for_old_databases()
    test_phase16_design_cancellation_marks_task_cancelled()
    test_phase16_long_dsrna_large_target_is_rejected_before_materializing_candidates()
    test_phase13_sgrna_validation_hits_use_amplicon_sequencing_action()
    print("delivery-readiness tests passed")
