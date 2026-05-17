"""
结果导出器
支持 CSV / Excel / FASTA 格式
"""

import csv
from pathlib import Path
from typing import List, Dict


class ResultExporter:
    """结果导出器"""

    def _sgrna_offtarget_count_fields(self, result: Dict) -> Dict:
        summary = ((result.get("off_target") or {}).get("summary") or {})
        ot_counts = summary.get("mismatch_counts") or {}
        pot_counts = summary.get("pot_mismatch_counts") or {}
        fields = {}
        for bucket in ["0M", "1M", "2M", "3M", "4M", "5M"]:
            fields[f"sgrna_ot_{bucket}"] = ot_counts.get(bucket, "")
            fields[f"sgrna_pot_{bucket}"] = pot_counts.get(bucket, "")
        fields.update({
            "sgrna_total_ot": summary.get("total_ot", ""),
            "sgrna_total_pot": summary.get("total_pot", ""),
            "sgrna_nrg_pam_hits": summary.get("nrg_pam_hits", ""),
            "sgrna_nag_pam_hits": summary.get("nag_pam_hits", summary.get("alternative_pam_hits", "")),
            "sgrna_risk_evaluation": summary.get("sgrnacas9_risk_evaluation", ""),
        })
        return fields

    def _target_reason_text(self, top_targets: List[Dict]) -> str:
        parts = []
        for target in top_targets:
            target_id = str(target.get("target_id", "")).strip()
            if not target_id:
                continue
            risk_score = target.get("risk_score", "")
            reasons = [str(reason) for reason in target.get("reasons", []) if reason]
            details = []
            if risk_score != "":
                details.append(f"score {risk_score}")
            if reasons:
                details.append(", ".join(reasons[:3]))
            parts.append(f"{target_id} ({'; '.join(details)})" if details else target_id)
        return "; ".join(parts)

    def _risk_fields(self, result: Dict) -> Dict:
        off_target = result.get("off_target") or {}
        if not off_target and ("risk_level" in result or "risk_score" in result):
            off_target = {
                "risk_level": result.get("risk_level", "low"),
                "risk_score": result.get("risk_score", 0),
                "top_targets": [
                    {"target_id": item.strip()}
                    for item in str(result.get("top_risk_targets", "")).split(";")
                    if item.strip()
                ],
                "validation_direction": result.get("validation_direction", ""),
            }
        top_targets = off_target.get("top_targets") or []
        target_text = "; ".join(
            f"{target.get('target_id', '')}:{target.get('risk_score', '')}"
            for target in top_targets
        )
        rnaup = result.get("rnaup") or {}
        rnaup_details = rnaup.get("details") or {}
        return {
            "recommendation_score": result.get("recommendation_score", result.get("consensus_score", "")),
            "cluster_id": result.get("cluster_id", ""),
            "cluster_size": result.get("cluster_size", 1),
            "alternative_count": result.get("alternative_count", 0),
            "cluster_span": result.get("cluster_span", result.get("position", "")),
            "off_target_risk": off_target.get("risk_level", "low"),
            "risk_score": off_target.get("risk_score", 0),
            "top_risk_targets": target_text,
            "top_offtarget_genes": self._target_reason_text(top_targets),
            "validation_direction": off_target.get("validation_direction", ""),
            "rnaup_dg": rnaup.get("dg", ""),
            "rnaup_method": rnaup_details.get("method", ""),
        }

    def _explanation_summary(self, result: Dict) -> str:
        explanation = result.get("explanation") or {}
        if isinstance(explanation, dict):
            return explanation.get("summary", "")
        return str(explanation)

    def _validation_text(self, result: Dict) -> str:
        hits = result.get("validation_hits") or []
        parts = []
        for hit in hits[:5]:
            parts.append(
                f"{hit.get('target_id', '')}:"
                f"{hit.get('match_type', '')}:"
                f"longest={hit.get('longest_contiguous_match', '')}:"
                f"{hit.get('validation_action', '')}"
            )
        return "; ".join(parts)

    def _primer_fields(self, result: Dict) -> Dict:
        primers = result.get("primers") or {}
        return {
            "product_length": primers.get("product_length", ""),
            "forward_primer": (primers.get("forward_primer") or {}).get("sequence", ""),
            "reverse_primer": (primers.get("reverse_primer") or {}).get("sequence", ""),
            "t7_forward_primer": (primers.get("t7_forward_primer") or {}).get("sequence", ""),
            "t7_reverse_primer": (primers.get("t7_reverse_primer") or {}).get("sequence", ""),
            "sgrna_forward_oligo": ((primers.get("sgrna_cloning_oligos") or {}).get("px330_forward_oligo") or {}).get("sequence", ""),
            "sgrna_reverse_oligo": ((primers.get("sgrna_cloning_oligos") or {}).get("px330_reverse_oligo") or {}).get("sequence", ""),
            "sgrna_custom_forward_oligo": ((primers.get("sgrna_cloning_oligos") or {}).get("custom_forward_oligo") or {}).get("sequence", ""),
            "sgrna_custom_reverse_oligo": ((primers.get("sgrna_cloning_oligos") or {}).get("custom_reverse_oligo") or {}).get("sequence", ""),
            "genotyping_forward_primer": (((primers.get("genotyping_primers") or {}).get("forward_primer") or {}).get("sequence", "")),
            "genotyping_reverse_primer": (((primers.get("genotyping_primers") or {}).get("reverse_primer") or {}).get("sequence", "")),
        }

    def _sgrna_fields(self, result: Dict) -> Dict:
        sgrna = result.get("sgrna") or {}
        input_advice = sgrna.get("input_advice") or []
        if isinstance(input_advice, list):
            input_advice = "; ".join(str(item) for item in input_advice if item)
        return {
            "sgrna_spacer_dna": sgrna.get("spacer_dna", ""),
            "sgrna_guide_rna": sgrna.get("guide_rna", result.get("sequence", "") if sgrna else ""),
            "sgrna_pam": sgrna.get("pam", ""),
            "sgrna_genomic_pam": sgrna.get("genomic_pam", sgrna.get("pam", "")),
            "sgrna_strand": sgrna.get("strand", ""),
            "sgrna_cut_site": sgrna.get("cut_site", ""),
            "sgrna_gc_percent": sgrna.get("gc_percent", ""),
            "sgrna_on_target_score": sgrna.get("on_target_score", ""),
            "sgrna_source_position_start": sgrna.get("source_position_start", ""),
            "sgrna_source_position_end": sgrna.get("source_position_end", ""),
            "sgrna_source_cut_site": sgrna.get("source_cut_site", ""),
            "sgrna_cds_region": sgrna.get("cds_region", ""),
            "sgrna_cds_position_percent": sgrna.get("cds_position_percent", ""),
            "sgrna_cds_priority_bonus": sgrna.get("cds_priority_bonus", ""),
            "sgrna_design_input_type": sgrna.get("design_input_type", ""),
            "sgrna_cds_source_start": sgrna.get("cds_source_start", ""),
            "sgrna_cds_source_end": sgrna.get("cds_source_end", ""),
            "sgrna_input_advice": input_advice,
        }

    def export_csv(self, results: List[Dict], filepath: str):
        """导出为 CSV"""
        if not results:
            return

        filepath = Path(filepath)
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            # 确定字段
            fieldnames = [
                "rank",
                "sequence",
                "position",
                "sgrna_spacer_dna",
                "sgrna_guide_rna",
                "sgrna_pam",
                "sgrna_genomic_pam",
                "sgrna_strand",
                "sgrna_cut_site",
                "sgrna_gc_percent",
                "sgrna_on_target_score",
                "sgrna_source_position_start",
                "sgrna_source_position_end",
                "sgrna_source_cut_site",
                "sgrna_cds_region",
                "sgrna_cds_position_percent",
                "sgrna_cds_priority_bonus",
                "sgrna_design_input_type",
                "sgrna_cds_source_start",
                "sgrna_cds_source_end",
                "sgrna_input_advice",
                "consensus_score",
                "recommendation_score",
                "cluster_id",
                "cluster_size",
                "alternative_count",
                "cluster_span",
                "off_target_risk",
                "risk_score",
                "sgrna_ot_0M",
                "sgrna_ot_1M",
                "sgrna_ot_2M",
                "sgrna_ot_3M",
                "sgrna_ot_4M",
                "sgrna_ot_5M",
                "sgrna_total_ot",
                "sgrna_pot_0M",
                "sgrna_pot_1M",
                "sgrna_pot_2M",
                "sgrna_pot_3M",
                "sgrna_pot_4M",
                "sgrna_pot_5M",
                "sgrna_total_pot",
                "sgrna_nrg_pam_hits",
                "sgrna_nag_pam_hits",
                "sgrna_risk_evaluation",
                "top_risk_targets",
                "top_offtarget_genes",
                "validation_direction",
                "rnaup_dg",
                "rnaup_method",
                "decision_summary",
                "validation_hits",
                "product_length",
                "forward_primer",
                "reverse_primer",
                "t7_forward_primer",
                "t7_reverse_primer",
                "sgrna_forward_oligo",
                "sgrna_reverse_oligo",
                "sgrna_custom_forward_oligo",
                "sgrna_custom_reverse_oligo",
                "genotyping_forward_primer",
                "genotyping_reverse_primer",
                "passed",
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for r in results:
                row = {
                    "rank": r.get("rank", ""),
                    "sequence": r.get("sequence", r.get("candidate_seq", "")),
                    "position": r.get("position", f"{r.get('position_start', '')}-{r.get('position_end', '')}"),
                    **self._sgrna_fields(r),
                    "consensus_score": r.get("consensus_score", ""),
                    **self._risk_fields(r),
                    **self._sgrna_offtarget_count_fields(r),
                    "decision_summary": self._explanation_summary(r),
                    "validation_hits": self._validation_text(r),
                    **self._primer_fields(r),
                    "passed": "Yes" if r.get("passed", False) else "No",
                }
                writer.writerow(row)

    def export_fasta(self, results: List[Dict], filepath: str):
        """导出为 FASTA"""
        filepath = Path(filepath)
        with open(filepath, "w", encoding="utf-8") as f:
            for r in results:
                seq = r.get("sequence", r.get("candidate_seq", ""))
                rank = r.get("rank", 0)
                score = r.get("consensus_score", 0)
                risk = self._risk_fields(r)
                f.write(
                    f">candidate_{rank}_score_{score:.2f}"
                    f"_clusterSize_{risk['cluster_size']}"
                    f"_risk_{risk['off_target_risk']}_riskScore_{risk['risk_score']}\n"
                )
                f.write(f"{seq}\n")

    def export_excel(self, results: List[Dict], filepath: str):
        """导出为 Excel"""
        try:
            import pandas as pd

            data = []
            for r in results:
                sgrna_fields = self._sgrna_fields(r)
                risk_fields = self._risk_fields(r)
                primer_fields = self._primer_fields(r)
                sgrna_counts = self._sgrna_offtarget_count_fields(r)
                data.append({
                    "Rank": r.get("rank", ""),
                    "Sequence": r.get("sequence", r.get("candidate_seq", "")),
                    "Position": r.get("position", f"{r.get('position_start', '')}-{r.get('position_end', '')}"),
                    "sgRNA Spacer DNA": sgrna_fields["sgrna_spacer_dna"],
                    "sgRNA Guide RNA": sgrna_fields["sgrna_guide_rna"],
                    "sgRNA PAM": sgrna_fields["sgrna_pam"],
                    "sgRNA Genomic PAM": sgrna_fields["sgrna_genomic_pam"],
                    "sgRNA Strand": sgrna_fields["sgrna_strand"],
                    "sgRNA Cut Site": sgrna_fields["sgrna_cut_site"],
                    "sgRNA GC %": sgrna_fields["sgrna_gc_percent"],
                    "sgRNA On-target Score": sgrna_fields["sgrna_on_target_score"],
                    "sgRNA Source Start": sgrna_fields["sgrna_source_position_start"],
                    "sgRNA Source End": sgrna_fields["sgrna_source_position_end"],
                    "sgRNA Source Cut Site": sgrna_fields["sgrna_source_cut_site"],
                    "sgRNA CDS Region": sgrna_fields["sgrna_cds_region"],
                    "sgRNA CDS Position %": sgrna_fields["sgrna_cds_position_percent"],
                    "sgRNA CDS Priority Bonus": sgrna_fields["sgrna_cds_priority_bonus"],
                    "sgRNA Design Input Type": sgrna_fields["sgrna_design_input_type"],
                    "sgRNA CDS Source Start": sgrna_fields["sgrna_cds_source_start"],
                    "sgRNA CDS Source End": sgrna_fields["sgrna_cds_source_end"],
                    "sgRNA Input Advice": sgrna_fields["sgrna_input_advice"],
                    "Consensus Score": r.get("consensus_score", ""),
                    "Recommendation Score": risk_fields["recommendation_score"],
                    "Cluster ID": risk_fields["cluster_id"],
                    "Cluster Size": risk_fields["cluster_size"],
                    "Alternative Count": risk_fields["alternative_count"],
                    "Cluster Span": risk_fields["cluster_span"],
                    "Off-target Risk": risk_fields["off_target_risk"],
                    "Risk Score": risk_fields["risk_score"],
                    "sgRNA OT 0M": sgrna_counts["sgrna_ot_0M"],
                    "sgRNA OT 1M": sgrna_counts["sgrna_ot_1M"],
                    "sgRNA OT 2M": sgrna_counts["sgrna_ot_2M"],
                    "sgRNA OT 3M": sgrna_counts["sgrna_ot_3M"],
                    "sgRNA OT 4M": sgrna_counts["sgrna_ot_4M"],
                    "sgRNA OT 5M": sgrna_counts["sgrna_ot_5M"],
                    "sgRNA Total OT": sgrna_counts["sgrna_total_ot"],
                    "sgRNA POT 0M": sgrna_counts["sgrna_pot_0M"],
                    "sgRNA POT 1M": sgrna_counts["sgrna_pot_1M"],
                    "sgRNA POT 2M": sgrna_counts["sgrna_pot_2M"],
                    "sgRNA POT 3M": sgrna_counts["sgrna_pot_3M"],
                    "sgRNA POT 4M": sgrna_counts["sgrna_pot_4M"],
                    "sgRNA POT 5M": sgrna_counts["sgrna_pot_5M"],
                    "sgRNA Total POT": sgrna_counts["sgrna_total_pot"],
                    "sgRNA NRG PAM Hits": sgrna_counts["sgrna_nrg_pam_hits"],
                    "sgRNA NAG PAM Hits": sgrna_counts["sgrna_nag_pam_hits"],
                    "sgRNA Risk Evaluation": sgrna_counts["sgrna_risk_evaluation"],
                    "Top Risk Targets": risk_fields["top_risk_targets"],
                    "Top Off-target Genes": risk_fields["top_offtarget_genes"],
                    "Validation Direction": risk_fields["validation_direction"],
                    "RNAup ΔG": risk_fields["rnaup_dg"],
                    "RNAup Method": risk_fields["rnaup_method"],
                    "Decision Summary": self._explanation_summary(r),
                    "Validation Hits": self._validation_text(r),
                    "Product Length": primer_fields["product_length"],
                    "Forward Primer": primer_fields["forward_primer"],
                    "Reverse Primer": primer_fields["reverse_primer"],
                    "T7 Forward Primer": primer_fields["t7_forward_primer"],
                    "T7 Reverse Primer": primer_fields["t7_reverse_primer"],
                    "sgRNA Forward Oligo": primer_fields["sgrna_forward_oligo"],
                    "sgRNA Reverse Oligo": primer_fields["sgrna_reverse_oligo"],
                    "sgRNA Custom Forward Oligo": primer_fields["sgrna_custom_forward_oligo"],
                    "sgRNA Custom Reverse Oligo": primer_fields["sgrna_custom_reverse_oligo"],
                    "Genotyping Forward Primer": primer_fields["genotyping_forward_primer"],
                    "Genotyping Reverse Primer": primer_fields["genotyping_reverse_primer"],
                    "Passed": "Yes" if r.get("passed", False) else "No",
                })

            df = pd.DataFrame(data)
            df.to_excel(filepath, index=False)
        except ImportError:
            raise ImportError("pandas required for Excel export. Install with: pip install openpyxl")

    def export_validation_report(self, results: List[Dict], filepath: str):
        """Export a multi-sheet experiment validation report."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl required for validation report export") from exc

        workbook = openpyxl.Workbook()
        recommendations = workbook.active
        recommendations.title = "Recommendations"
        recommendations.append([
            "Rank", "Sequence", "Position", "Consensus Score", "Recommendation Score",
            "sgRNA PAM", "sgRNA Genomic PAM", "sgRNA Strand", "sgRNA Cut Site",
            "sgRNA Source Start", "sgRNA Source End", "sgRNA CDS Region", "sgRNA CDS Position %",
            "Risk", "Risk Score", "sgRNA Total OT", "sgRNA Total POT",
            "sgRNA NRG PAM Hits", "sgRNA NAG PAM Hits", "sgRNA Risk Evaluation",
            "Top Off-target Genes", "Decision Summary", "Validation Direction", "Passed",
        ])

        validation_sheet = workbook.create_sheet("OffTarget Validation")
        validation_sheet.append([
            "Rank", "Candidate", "Target", "Risk Score", "Match Type", "Target Position",
            "Reasons", "Longest Match", "Mismatches", "Seed Match", "Validation Action",
            "Query Fragment", "Target Fragment",
        ])

        primer_sheet = workbook.create_sheet("Primers")
        primer_sheet.append([
            "Rank", "Product Length", "Forward Primer", "Reverse Primer",
            "T7 Forward Primer", "T7 Reverse Primer", "sgRNA Forward Oligo",
            "sgRNA Reverse Oligo", "sgRNA Custom Forward Oligo", "sgRNA Custom Reverse Oligo",
            "Genotyping Forward Primer", "Genotyping Reverse Primer", "Notes",
        ])

        methods = workbook.create_sheet("Methods")
        methods.append(["Section", "Text"])
        methods.append(["Ranking", "Recommendation score combines efficacy score and off-target risk penalty."])
        methods.append(["Dicer", "Long dsRNA product pools are heuristic Dicer-product simulations for relative ranking, not quantitative knockdown predictions."])
        methods.append(["Off-target", "RNAi risk uses 27bp/20bp/16bp continuous matches and 7nt seed matches; sgRNA risk uses SpCas9 NRG PAM-adjacent <=5 mismatch spacer hits, with separate 12 nt seed/POT counts."])
        methods.append(["sgRNA input", "sgRNA design prefers CDS input. If mRNA/cDNA is supplied, the software attempts longest ATG-to-stop CDS inference and prioritizes front-CDS guides; source coordinates are reported against the original input."])
        methods.append(["RNAup", "RNAup CLI values are reported only when method is RNAup-cli; otherwise fallback is explicitly labeled."])
        methods.append(["Primers", "T7 primers are PCR primers prefixed with the T7 promoter sequence; verify amplicon uniqueness and primer structure before ordering."])

        for r in results:
            risk = self._risk_fields(r)
            sgrna = self._sgrna_fields(r)
            sgrna_counts = self._sgrna_offtarget_count_fields(r)
            recommendations.append([
                r.get("rank", ""),
                r.get("sequence", r.get("candidate_seq", "")),
                r.get("position", f"{r.get('position_start', '')}-{r.get('position_end', '')}"),
                r.get("consensus_score", ""),
                risk["recommendation_score"],
                sgrna["sgrna_pam"],
                sgrna["sgrna_genomic_pam"],
                sgrna["sgrna_strand"],
                sgrna["sgrna_cut_site"],
                sgrna["sgrna_source_position_start"],
                sgrna["sgrna_source_position_end"],
                sgrna["sgrna_cds_region"],
                sgrna["sgrna_cds_position_percent"],
                risk["off_target_risk"],
                risk["risk_score"],
                sgrna_counts["sgrna_total_ot"],
                sgrna_counts["sgrna_total_pot"],
                sgrna_counts["sgrna_nrg_pam_hits"],
                sgrna_counts["sgrna_nag_pam_hits"],
                sgrna_counts["sgrna_risk_evaluation"],
                risk["top_offtarget_genes"],
                self._explanation_summary(r),
                risk["validation_direction"],
                "Yes" if r.get("passed", False) else "No",
            ])

            for hit in r.get("validation_hits") or []:
                validation_sheet.append([
                    r.get("rank", ""),
                    r.get("sequence", r.get("candidate_seq", "")),
                    hit.get("target_id", ""),
                    hit.get("risk_score", ""),
                    hit.get("match_type", ""),
                    hit.get("target_position", ""),
                    ", ".join(str(reason) for reason in hit.get("reasons", []) if reason),
                    hit.get("longest_contiguous_match", ""),
                    hit.get("mismatch_count", ""),
                    "Yes" if hit.get("seed_match") else "No",
                    hit.get("validation_action", ""),
                    hit.get("query_fragment", ""),
                    hit.get("target_fragment", ""),
                ])

            primers = r.get("primers") or {}
            if primers:
                cloning = primers.get("sgrna_cloning_oligos") or {}
                genotyping = primers.get("genotyping_primers") or {}
                primer_sheet.append([
                    r.get("rank", ""),
                    primers.get("product_length", ""),
                    (primers.get("forward_primer") or {}).get("sequence", ""),
                    (primers.get("reverse_primer") or {}).get("sequence", ""),
                    (primers.get("t7_forward_primer") or {}).get("sequence", ""),
                    (primers.get("t7_reverse_primer") or {}).get("sequence", ""),
                    (cloning.get("px330_forward_oligo") or {}).get("sequence", ""),
                    (cloning.get("px330_reverse_oligo") or {}).get("sequence", ""),
                    (cloning.get("custom_forward_oligo") or {}).get("sequence", ""),
                    (cloning.get("custom_reverse_oligo") or {}).get("sequence", ""),
                    (genotyping.get("forward_primer") or {}).get("sequence", ""),
                    (genotyping.get("reverse_primer") or {}).get("sequence", ""),
                    primers.get("notes", cloning.get("notes", genotyping.get("notes", ""))),
                ])

        workbook.save(filepath)

    def export_primer_order_csv(self, results: List[Dict], filepath: str):
        """Export primer rows suitable for copy/paste into an ordering sheet."""
        rows = []
        for r in results:
            primers = r.get("primers") or {}
            for key in ["forward_primer", "reverse_primer", "t7_forward_primer", "t7_reverse_primer"]:
                primer = primers.get(key) or {}
                if primer.get("sequence"):
                    rows.append({
                        "candidate_rank": r.get("rank", ""),
                        "primer_name": f"candidate_{r.get('rank', '')}_{key}",
                        "sequence": primer["sequence"],
                        "length": primer.get("length", ""),
                        "tm": primer.get("tm", ""),
                        "gc_percent": primer.get("gc_percent", ""),
                    })
            cloning = primers.get("sgrna_cloning_oligos") or {}
            for key in ["px330_forward_oligo", "px330_reverse_oligo", "custom_forward_oligo", "custom_reverse_oligo"]:
                oligo = cloning.get(key) or {}
                if oligo.get("sequence"):
                    rows.append({
                        "candidate_rank": r.get("rank", ""),
                        "primer_name": f"candidate_{r.get('rank', '')}_{key}",
                        "sequence": oligo["sequence"],
                        "length": oligo.get("length", ""),
                        "tm": "",
                        "gc_percent": "",
                    })
            genotyping = primers.get("genotyping_primers") or {}
            for key in ["forward_primer", "reverse_primer"]:
                primer = genotyping.get(key) or {}
                if primer.get("sequence"):
                    rows.append({
                        "candidate_rank": r.get("rank", ""),
                        "primer_name": f"candidate_{r.get('rank', '')}_genotyping_{key}",
                        "sequence": primer["sequence"],
                        "length": primer.get("length", ""),
                        "tm": primer.get("tm", ""),
                        "gc_percent": primer.get("gc_percent", ""),
                    })
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["candidate_rank", "primer_name", "sequence", "length", "tm", "gc_percent"],
            )
            writer.writeheader()
            writer.writerows(rows)
